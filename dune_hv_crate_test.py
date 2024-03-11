import pyvisa
import sys
import json
import os
import time
import openpyxl
from datetime import datetime
from keysight_daq970a import Keysight970A
from rigol_dp832a import RigolDP832A
from caen_r8033dm_wrapper import CAENR8033DM_WRAPPER

import csv
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator
import numpy as np
from scipy.optimize import curve_fit

class LDOmeasure:
    def __init__(self, config_file, name = None):
        self.prefix = "DUNE HV Crate Tester"
        print(f"{self.prefix} --> Welcome to the DUNE HV crate production testing script")
        with open(config_file, "r") as jsonfile:
            self.json_data = json.load(jsonfile)
        self.rm = pyvisa.ResourceManager('@py')

        #Initialize all instruments first so that you don't waste time with input if something is not connected
        self.c = CAENR8033DM_WRAPPER(self.json_data)
        self.k = Keysight970A(self.rm, self.json_data)

        #Since there are 2 Rigols, set them up here so they know what channels they have
        #And if the test sequence calls the wrong one, it'll throw an error
        self.r0 = RigolDP832A(self.rm, self.json_data, 0)
        self.r0.setup_fan()
        self.r0.setup_heater_supply()
        self.r0.setup_heater_switch()

        self.r1 = RigolDP832A(self.rm, self.json_data, 1)
        self.r1.setup_hvpullup()
        self.r1.setup_hvpullup2()
        self.r1.setup_fanread()

        #Now we can get the input for the name of the test
        if (name):
            self.test_name = name
        else:
            self.test_name = input("Input the test name:\n")

        self.rounding_factor = self.json_data["rounding_factor"]
        #The datastore is the eventual output JSON file that will be written after the test
        #Want to also include what the inputs for this particular test was
        self.datastore = {}
        self.datastore['input_params'] = self.json_data
        self.datastore['test_name'] = self.test_name
        self.start_time = datetime.now()
        self.datastore['start_time'] = self.start_time
        self.initialize_spreadsheet()

        self.fan_test_result = False
        self.heat_test_result = False
        self.hv_test_result = False

        self.datastore['Tests'] = {}

        self.fan_test()
        self.heater_test()
        self.hv_test()

        if (self.fan_test_result and self.heat_test_result and self.hv_test_result):
            self.ws.cell(row=self.row, column=1, value=self.test_name).style = "pass"
            self.datastore['overall'] = "Pass"
        else:
            self.ws.cell(row=self.row, column=1, value=self.test_name).style = "fail"
            self.datastore['overall'] = "Fail"
        self.wb.save(self.path_to_spreadsheet)

        end_time = datetime.now()
        test_time = end_time - self.start_time
        self.datastore['end_time'] = end_time
        self.datastore['test_time'] = test_time

        with open(self.json_output_file, 'w', encoding='utf-8') as f:
            json.dump(self.datastore, f, ensure_ascii=False, indent=4, default=str)

        print(f"{self.prefix} --> Test complete")
        self.make_hv_plots()

    #Looks to see if a main spreadsheet of all results exists. If it does, open it and find the next row to write these results to
    #If not, it creates the spreadsheet with the proper headers and formatting
    def initialize_spreadsheet(self):
        #In the config JSON I give the user the ability to choose the absolute path to dump all this stuff into, or make it relative to this Python script
        if (self.json_data["relative"] == "True"):
            output_path = os.path.abspath(self.json_data["output_directory"])
        else:
            output_path = os.path.normpath(self.json_data["output_directory"])

        self.path_to_spreadsheet = os.path.join(output_path, f"{self.json_data['output_file']}")
        self.datastore['spreadsheet_path'] = self.path_to_spreadsheet

        json_date = datetime.today().strftime('%Y%m%d%H%M%S')
        #json_date = "20240125152313"
        os.makedirs(os.path.join(output_path, json_date))
        self.results_path = os.path.join(output_path, json_date)
        self.json_output_file = os.path.join(self.results_path, f"{json_date}_{self.test_name}.json")
        self.datastore['json_path'] = self.json_output_file

        self.hv_cols = 8

        if (os.path.isfile(self.path_to_spreadsheet)):
            self.wb = openpyxl.load_workbook(filename = self.path_to_spreadsheet)
            self.ws = self.wb.active
            self.row = self.ws.max_row + 1
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = self.json_data["sheet_name"]

            #Once styles are added, they are saved even if you close and reopen the spreadsheet
            top_style = openpyxl.styles.NamedStyle(name="top")
            top_style.font = openpyxl.styles.Font(bold=True, name='Calibri')
            bd = openpyxl.styles.Side(style='thin')
            top_style.border = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
            self.wb.add_named_style(top_style)

            self.fail_style = openpyxl.styles.NamedStyle(name="fail")
            self.fail_style.font = openpyxl.styles.Font(color="FF0000", name='Calibri')
            self.wb.add_named_style(self.fail_style)

            self.pass_style = openpyxl.styles.NamedStyle(name="pass")
            self.pass_style.font = openpyxl.styles.Font(color="009900", name='Calibri')
            self.wb.add_named_style(self.pass_style)

            #Yes I'm using magic numbers, but I need to build the spreadsheet in a specific way, hopefully it's only done once here
            self.ws.cell(row=2, column=1, value="Test Name").style = top_style
            self.ws.cell(row=2, column=2, value="Date").style = top_style
            self.ws.cell(row=2, column=3, value="Time").style = top_style
            self.ws.cell(row = 1, column = 4, value="Fan Test - V/I for power supplying all 4 fans, RD signal outputs").style = top_style
            self.ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=9)
            self.ws.cell(row=2, column=4, value="Supply Voltage").style = top_style
            self.ws.cell(row=2, column=5, value="Supply Current").style = top_style
            for i in range(1,5):
                self.ws.cell(row=2, column=5+i, value=f"Fan {i} RD").style = top_style
                self.ws.cell(row=2, column=9+i, value=f"TC{i}_Resistance").style = top_style
                self.ws.cell(row=2, column=13+i, value=f"TC{i}_Temp_Rise").style = top_style

            self.ws.cell(row=1, column=10, value="Heater Test - Results for each heating element and temperature rise after heating time").style = top_style
            self.ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=17)

            self.ws.cell(row=1, column=18, value="HV Test - Results for each configuration. Resistance in Mega Ohms, time constant is tau (seconds) in a*e^(-tau * t)+c ").style = top_style
            self.ws.merge_cells(start_row=1, start_column=18, end_row=1, end_column=21+(7*self.hv_cols))
            for i in range(8):
                self.ws.cell(row=2, column=18+(i*self.hv_cols), value=f"Ch{i}+ Open Res").style = top_style
                self.ws.cell(row=2, column=19+(i*self.hv_cols), value=f"Ch{i}+ Open Fit").style = top_style
                self.ws.cell(row=2, column=20+(i*self.hv_cols), value=f"Ch{i}+ 10k Res").style = top_style
                self.ws.cell(row=2, column=21+(i*self.hv_cols), value=f"Ch{i}+ 10k Fit").style = top_style
                self.ws.cell(row=2, column=22+(i*self.hv_cols), value=f"Ch{i}- Open Res").style = top_style
                self.ws.cell(row=2, column=23+(i*self.hv_cols), value=f"Ch{i}- Open Fit").style = top_style
                self.ws.cell(row=2, column=24+(i*self.hv_cols), value=f"Ch{i}- 10k Res").style = top_style
                self.ws.cell(row=2, column=25+(i*self.hv_cols), value=f"Ch{i}- 10k Fit").style = top_style

            #Expands each column to have the best width to fit everything
            column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(self.ws.max_column))
            for column_letter in column_letters:
                self.ws.column_dimensions[column_letter].bestFit = True
            self.ws.freeze_panes = self.ws.cell(row=3, column=2)
            self.wb.save(self.path_to_spreadsheet)
            self.row = 3

        #In any case, start filling in the spreadsheet with initial test parameters we have right know
        #Test name starts in red so that if the test is incomplete it will show as a fail
        self.ws.cell(row=self.row, column=1, value=self.test_name).style = "fail"
        self.ws.cell(row=self.row, column=2, value=datetime.today().strftime('%m/%d/%Y'))
        self.ws.cell(row=self.row, column=3, value=datetime.today().strftime('%I:%M:%S %p'))
        self.wb.save(self.path_to_spreadsheet)

    def fan_test(self):
        #Fan test
        self.k.initialize_fan()
        self.r0.power("ON", "fan")
        self.r1.power("ON", "fanread")
        print(f"{self.prefix} --> Fans turned on, waiting {self.json_data['fan_wait']} seconds for the fans to reach steady state...")
        time.sleep(self.json_data['fan_wait'])
        fan_voltage = self.r0.get_voltage("fan")
        fan_current = self.r0.get_current("fan")
        fanread_voltage = self.r1.get_voltage("fanread")
        fanread_current = self.r1.get_current("fanread")
        fan_read_signal = self.k.measure_fan()
        self.r0.power("OFF", "fan")
        self.r1.power("OFF", "fanread")
        print(f"{self.prefix} --> Fans turned off")
        print(f"{self.prefix} --> Fan power supply was {fan_voltage}V and {fan_current}A")
        print(f"{self.prefix} --> Read signal for each fan was {fan_read_signal}")
        print(f"{self.prefix} --> Fan read pullup supply was {fanread_voltage}V and {fanread_current}A")

        self.fan_test_result = True
        if ((fan_voltage < self.json_data["fan_voltage_max"]) and (fan_voltage > self.json_data["fan_voltage_min"])):
            self.ws.cell(row=self.row, column=4, value=fan_voltage)
            self.datastore['Tests']['fan_voltage_test'] = "Pass"
        else:
            self.ws.cell(row=self.row, column=4, value=fan_voltage).style = "fail"
            self.datastore['Tests']['fan_voltage_test'] = "Fail"
            self.fan_test_result = False

        if ((fan_current < self.json_data["fan_current_max"]) and (fan_current > self.json_data["fan_current_min"])):
            self.ws.cell(row=self.row, column=5, value=fan_current)
            self.datastore['Tests']['fan_current_test'] = "Pass"
        else:
            self.ws.cell(row=self.row, column=5, value=fan_current).style = "fail"
            self.datastore['Tests']['fan_current_test'] = "Fail"
            self.fan_test_result = False

        for i in range(1,5):
            if ((fan_read_signal[i] < self.json_data["fan_read_max"]) and (fan_read_signal[i] > self.json_data["fan_read_min"])):
                self.ws.cell(row=self.row, column=5+i, value=round(fan_read_signal[i], self.rounding_factor))
                self.datastore['Tests'][f'fan_signal_test_{i}'] = "Pass"
            else:
                self.ws.cell(row=self.row, column=5+i, value=round(fan_read_signal[i], self.rounding_factor)).style = "fail"
                self.fan_test_result = False

        self.datastore['fan_voltage'] = fan_voltage
        self.datastore['fan_current'] = fan_current
        self.datastore['fanread_voltage'] = fanread_voltage
        self.datastore['fanread_current'] = fanread_current
        self.datastore['fan_read_signal'] = fan_read_signal

    def heater_test(self):
        #Heater test
        #First measure resistance of heating element with no power connected
        self.k.initialize_resistance()
        heater_resistance = self.k.measure_resistance()
        print(f"{self.prefix} --> Heating element resistances are {heater_resistance}")

        self.heat_test_result = True
        for i in range(1,5):
            if ((heater_resistance[i] < self.json_data["heating_element_max"]) and (heater_resistance[i] > self.json_data["heating_element_min"])):
                self.ws.cell(row=self.row, column=9+i, value=round(heater_resistance[i], self.rounding_factor))
                self.datastore['Tests'][f'heating_element_test_{i}'] = "Pass"
            else:
                self.ws.cell(row=self.row, column=9+i, value=round(heater_resistance[i], self.rounding_factor)).style = "fail"
                self.datastore['Tests'][f'heating_element_test_{i}'] = "Fail"
                self.heat_test_result = False

        self.datastore['heater_resistance'] = heater_resistance

        #Then prepare RTD and switch relay to connect power
        self.k.initialize_rtd()
        temp1 = self.k.measure_rtd()
        self.r0.power("ON", "heat_supply")
        self.r0.power("ON", "heat_switch")
        print(f"{self.prefix} --> Heat turned on, waiting {self.json_data['heat_wait']} seconds for the sensors to heat up...")
        time.sleep(self.json_data['heat_wait'])
        supply_voltage = self.r0.get_voltage("heat_supply")
        supply_current = self.r0.get_current("heat_supply")
        switch_voltage = self.r0.get_voltage("heat_switch")
        switch_current = self.r0.get_current("heat_switch")
        temp2 = self.k.measure_rtd()
        temp_rise = []
        temp_rise.append(temp2[1] - temp1[1])
        temp_rise.append(temp2[2] - temp1[2])
        temp_rise.append(temp2[3] - temp1[3])
        temp_rise.append(temp2[4] - temp1[4])

        self.r0.power("OFF", "heat_supply")
        self.r0.power("OFF", "heat_switch")
        print(f"{self.prefix} --> Heat turned off")
        print(f"{self.prefix} --> Heat power supply was {supply_voltage}V and {supply_current}A")
        print(f"{self.prefix} --> Heat power switch was {switch_voltage}V and {switch_current}A")
        print(f"{self.prefix} --> Original temperatures were {temp1}")
        print(f"{self.prefix} --> Temperatures after {self.json_data['heat_wait']} seconds were {temp2}, a rise of {temp_rise}")

        for i in range(4):
            if ((temp_rise[i] < self.json_data["temp_increase_max"]) and (temp_rise[i] > self.json_data["temp_increase_min"])):
                self.ws.cell(row=self.row, column=14+i, value=round(temp_rise[i], self.rounding_factor))
                self.datastore['Tests'][f'temperature_rise_test_{i}'] = "Pass"
            else:
                self.ws.cell(row=self.row, column=14+i, value=round(temp_rise[i], self.rounding_factor)).style = "fail"
                self.datastore['Tests'][f'temperature_rise_test_{i}'] = "Fail"
                self.heat_test_result = False

        self.datastore['heater_supply_voltage'] = supply_voltage
        self.datastore['heater_supply_current'] = supply_current
        self.datastore['heater_switch_voltage'] = switch_voltage
        self.datastore['heater_switch_current'] = switch_current

        self.datastore['temp1'] = temp1
        self.datastore['temp2'] = temp2
        self.datastore['temp_rise'] = temp_rise

    def hv_test(self):
        #HV Leakage Test
        hv_results = {}
        self.r1.power("ON", "hvpullup")
        self.r1.power("ON", "hvpullup2")
        self.hv_test_result = True
        for i in range(1):
            hv_results[i] = {}

            #Measure the ramp from 0 to positive voltage with open termination
            print(f"{self.prefix} --> Turning Channel {i} HV from 0 to {self.json_data['caenR8033DM_voltage']}V with open termination")
            self.k.set_relay(0, 1 << i)
            self.c.turn_on(i)
            print(f"{self.prefix} --> HV reached max value, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_stability_wait'])

            csv_name = f"{self.test_name}_ch{i}_pos_open_on.csv"
            self.record_hv_data(csv_name)
            fit = self.hv_curve_fit(csv_name, i)
            hv_results[i]["pos_open_fit"] = fit
            hv_results[i]["pos_open_V"] = self.c.get_voltage(i)
            hv_results[i]["pos_open_I"] = self.c.get_current(i)

            #Measure the ramp from positive voltage to 0 with open termination
            print(f"{self.prefix} --> Turning Channel {i} HV from {self.json_data['caenR8033DM_voltage']}V to 0 with open termination")
            self.c.turn_off(i)
            print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_stability_wait'])

            #self.record_hv_data(f"{self.test_name}_ch{i}_pos_open_off.csv")

            #Measure the ramp from 0 to positive voltage with 10k termination
            print(f"{self.prefix} --> Turning Channel {i} HV from 0 to {self.json_data['caenR8033DM_voltage']}V with 10k termination")
            self.k.set_relay(0, 0)
            self.c.turn_on(i)
            print(f"{self.prefix} --> HV reached max value, waiting {self.json_data['hv_termination_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_termination_wait'])

            csv_name = f"{self.test_name}_ch{i}_pos_10k_on.csv"
            self.record_hv_data(f"{self.test_name}_ch{i}_pos_10k_on.csv")
            fit = self.hv_curve_fit(csv_name, i, term = True)
            hv_results[i]["pos_term_fit"] = fit
            hv_results[i]["pos_term_V"] = self.c.get_voltage(i)
            hv_results[i]["pos_term_I"] = self.c.get_current(i)

            #Measure the ramp from positive voltage to 0 with 10k termination
            print(f"{self.prefix} --> Turning Channel {i} HV from {self.json_data['caenR8033DM_voltage']}V to 0 with 10k termination")
            self.c.turn_off(i)
            print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_stability_wait'])

            #self.record_hv_data(f"{self.test_name}_ch{i}_pos_10k_off.csv")

            #Measure the ramp from 0 to negative voltage with open termination
            print(f"{self.prefix} --> Turning Channel {i} HV from 0 to -{self.json_data['caenR8033DM_voltage']}V with open termination")
            self.k.set_relay(1 << i, 1 << i)
            self.c.turn_on(i+8)
            print(f"{self.prefix} --> HV reached max value, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_stability_wait'])

            csv_name = f"{self.test_name}_ch{i}_neg_open_on.csv"
            self.record_hv_data(csv_name)
            fit = self.hv_curve_fit(csv_name, i+8)
            hv_results[i]["neg_open_fit"] = fit
            hv_results[i]["neg_open_V"] = self.c.get_voltage(i+8)
            hv_results[i]["neg_open_I"] = self.c.get_current(i+8)

            #Measure the ramp from negative voltage to 0 with open termination
            print(f"{self.prefix} --> Turning Channel {i} HV from -{self.json_data['caenR8033DM_voltage']}V to 0 with open termination")
            self.c.turn_off(i+8)
            print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_stability_wait'])

            #self.record_hv_data(f"{self.test_name}_ch{i}_neg_open_off.csv")


            #Measure the ramp from 0 to negative voltage with 10k termination
            print(f"{self.prefix} --> Turning Channel {i} HV from 0 to -{self.json_data['caenR8033DM_voltage']}V with 10k termination")
            self.k.set_relay(1 << i, 0)
            self.c.turn_on(i+8)
            print(f"{self.prefix} --> HV reached max value, waiting {self.json_data['hv_termination_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_termination_wait'])

            csv_name = f"{self.test_name}_ch{i}_neg_10k_on.csv"
            self.record_hv_data(f"{self.test_name}_ch{i}_neg_10k_on.csv")
            fit = self.hv_curve_fit(csv_name, i+8, term = True)
            hv_results[i]["neg_term_fit"] = fit
            hv_results[i]["neg_term_V"] = self.c.get_voltage(i+8)
            hv_results[i]["neg_term_I"] = self.c.get_current(i+8)

            #Measure the ramp from 0 to negative voltage with 10k termination
            print(f"{self.prefix} --> Turning Channel {i} HV from -{self.json_data['caenR8033DM_voltage']}V to 0 with 10k termination")
            self.c.turn_off(i+8)
            print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            time.sleep(self.json_data['hv_stability_wait'])

            #self.record_hv_data(f"{self.test_name}_ch{i}_neg_10k_off.csv")

            self.r1.power("OFF", "hvpullup")
            self.r1.power("OFF", "hvpullup2")

            #Voltage is in volts, current is in microamps, R in Mohms
            try:
                hv_results[i]["pos_open_R"] = float(hv_results[i]["pos_open_V"])/float(hv_results[i]["pos_open_I"])
            except:
                hv_results[i]["pos_open_R"] = 0
            try:
                hv_results[i]["pos_term_R"] = float(hv_results[i]["pos_term_V"])/float(hv_results[i]["pos_term_I"])
            except:
                hv_results[i]["pos_term_R"] = 0
            try:
                hv_results[i]["neg_open_R"] = float(hv_results[i]["neg_open_V"])/float(hv_results[i]["neg_open_I"])
            except:
                hv_results[i]["neg_open_R"] = 0
            try:
                hv_results[i]["neg_term_R"] = float(hv_results[i]["neg_term_V"])/float(hv_results[i]["neg_term_I"])
            except:
                hv_results[i]["neg_term_R"] = 0

            print(f"{self.prefix} --> Channel {i} HV results are {hv_results[i]}")

            for num,j in enumerate(["pos_open_R", "pos_term_R", "neg_open_R", "neg_term_R"]):
                if (num % 2):
                    max_val = self.json_data["hv_resistance_term_max"]
                    min_val = self.json_data["hv_resistance_term_min"]
                else:
                    max_val = self.json_data["hv_resistance_open_max"]
                    min_val = self.json_data["hv_resistance_open_min"]

                if ((float(hv_results[i][j]) < max_val) and (float(hv_results[i][j]) > min_val)):
                    self.ws.cell(row=self.row, column=18+(i*self.hv_cols)+(num*2), value=round(float(hv_results[i][j]), self.rounding_factor))
                    self.datastore['Tests'][f'hv_test_ch{i}_{j}'] = "Pass"
                else:
                    self.ws.cell(row=self.row, column=18+(i*self.hv_cols)+(num*2), value=round(float(hv_results[i][j]), self.rounding_factor)).style = "fail"
                    self.datastore['Tests'][f'hv_test_ch{i}_{j}'] = "Fail"
                    self.hv_test_result = False

            for num,j in enumerate(["pos_open_fit", "pos_term_fit", "neg_open_fit", "neg_term_fit"]):
                if ((float(hv_results[i][j][0][1]) < self.json_data["hv_tau_max"]) and (float(hv_results[i][j][0][1]) > self.json_data["hv_tau_min"])):
                    self.ws.cell(row=self.row, column=19+(i*self.hv_cols)+(num*2), value=round(float(hv_results[i][j][0][1]), self.rounding_factor))
                    self.datastore['Tests'][f'hv_fit_test_ch{i}_{j}'] = "Pass"
                else:
                    self.ws.cell(row=self.row, column=19+(i*self.hv_cols)+(num*2), value=round(float(hv_results[i][j][0][1]), self.rounding_factor)).style = "fail"
                    self.datastore['Tests'][f'hv_fit_test_ch{i}_{j}'] = "Fail"
                    self.hv_test_result = False

            self.datastore[f'hv_ch{i}'] = {}
            for j in ["pos_open_V", "pos_open_I", "pos_open_R", "neg_open_V", "neg_open_I", "neg_open_R", "pos_open_fit", "neg_open_fit",
                      "pos_term_V", "pos_term_I", "pos_term_R", "neg_term_V", "neg_term_I", "neg_term_R", "pos_term_fit", "neg_term_fit"]:
                self.datastore[f'hv_ch{i}'][j] = hv_results[i][j]

    def record_hv_data(self, name):
        data = []
        cycle_start_time = time.time()
        prev_measurement = cycle_start_time - 1
        print(f"{self.prefix} --> Collecting data for {name} for {self.json_data['hv_minutes_duration']} minutes starting at {cycle_start_time}...")
        while (time.time() - cycle_start_time < (self.json_data['hv_minutes_duration'] * 60)):
            if (time.time() > prev_measurement + self.json_data['hv_seconds_interval']):
                #print(f"{self.prefix} --> Measurement taken at {time.time()}")
                prev_measurement = prev_measurement + self.json_data['hv_seconds_interval']
                datum = [datetime.now()]
                for i in range(16):
                    datum.append(self.c.get_voltage(i))
                    datum.append(self.c.get_current(i))
                data.append(datum)
        with open(os.path.join(self.results_path, name), 'w') as fp:
            csv_writer = csv.writer(fp, delimiter=',')
            csv_writer.writerows(data)
        #input("ok?")

    def hv_curve_fit(self, name, ch, term = False):
        ch_datetime = []
        ch_voltage = []
        ch_current = []
        with open(os.path.join(self.results_path, name), 'r', newline='') as csvfile:
            spamreader = csv.reader(csvfile, delimiter=',')
            for row in spamreader:
                ch_datetime.append(datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S.%f'))
                ch_voltage.append(float(row[1 + (ch*2)]))
                if (term):
                    ch_current.append(float(row[2 + (ch*2)])/1000)
                else:
                    ch_current.append(float(row[2 + (ch*2)]))

        #print(f"For channel {ch}, grabbed column {2 + (ch*2)} and got this data")
        #print(ch_current)
        first_time = ch_datetime[0]
        ch_timedelta = [i-first_time for i in ch_datetime]
        ch_time = [datetime(2024, 1, 1, 0, i.seconds//60%60, i.seconds%60, 0) for i in ch_timedelta]

        def exp_fit(x, a, b, c):
            y = a*np.exp(-b*x) + c
            return y

        first_timestamp = ch_time[0].timestamp()
        time_seconds = [dt.timestamp() - first_timestamp for dt in ch_time]
        try:
            fit = curve_fit(exp_fit, time_seconds, ch_current)
        except RuntimeError:
            fit = [[0,0]]
        return fit

    def make_hv_plots(self):
        ch0_pos_open_fit = self.datastore['hv_ch0']['pos_open_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_pos_open_on", "0 to 2kV, open termination", True, True, 0, ch0_pos_open_fit)
        # self.make_plot(f"{self.test_name}_ch0_pos_open_off", "2kV to 0, open termination", False, True)
        ch0_pos_term_fit = self.datastore['hv_ch0']['pos_term_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_pos_10k_on", "0 to 2kV, 10k termination", True, True, 0, ch0_pos_term_fit)
        #self.make_plot(f"{self.test_name}_ch0_pos_10k_off", "2kV to 0, 10k termination", False, True, 0)

        ch0_neg_open_fit = self.datastore['hv_ch0']['neg_open_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_neg_open_on", "0 to -2kV, open termination", True, False, 8, ch0_neg_open_fit)
        # self.make_plot(f"{self.test_name}_ch0_neg_open_off", "-2kV to 0, open termination", False, False)
        ch0_neg_term_fit = self.datastore['hv_ch0']['neg_term_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_neg_10k_on", "0 to -2kV, 10k termination", True, False, 8, ch0_neg_term_fit)
        #self.make_plot(f"{self.test_name}_ch0_neg_10k_off", "-2kV to 0, 10k termination", False, True, 8)

    def make_plot(self, filename, name, on, pos, ch, fit=None):
        ch1_time, ch1_voltage, ch1_current = self.get_ch_data(os.path.join(self.results_path, f"{filename}.csv"), ch)
        # self.make_plot(f"{self.test_name}_ch0_neg_10k_off", "-2kV to 0, 10k termination", False, False)

        fig = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig.add_subplot(1,1,1)

        ax.plot(ch1_time, ch1_current, label="Ch Current")
        self.format_plot(ax)

        ax2 = ax.twinx()
        ax2.plot(ch1_time, ch1_voltage, label="Ch Voltage", color="red")

        fig.suptitle((name), fontsize=36)

        ax.set_xlabel("Time (Minutes:Seconds)", fontsize=24)
        ax.set_ylabel("Current (uA)", fontsize=24)

        # ax.set_xlim([0,150])
        if (on):
            ax2.set_ylim([115,121])
        elif (not on and not pos):
            ax2.set_ylim([-1,1])
        ax2.set_ylabel("Voltage (V)", fontsize=24)
        self.format_plot(ax2)

        if (on and fit):
            textstr = r'$\tau=%.4f$' % (fit)
            props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
            ax.text(0.75, 0.75, textstr, transform=ax.transAxes, fontsize=24,
            verticalalignment='top', bbox=props)

        fig.legend(loc='lower left', prop={'size': 20}, ncol=2)
        fig.savefig(os.path.join(self.results_path, f"{filename}.png"))
        plt.close(fig)

    def get_ch_data(self, data_file, ch):
        ch1_datetime = []
        ch1_voltage = []
        ch1_current = []
        with open(data_file, 'r', newline='') as csvfile:
            spamreader = csv.reader(csvfile, delimiter=',')
            for row in spamreader:
                ch1_datetime.append(datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S.%f'))
                ch1_voltage.append(float(row[1+(ch*2)]))
                ch1_current.append(float(row[2+(ch*2)]))

        first_time = ch1_datetime[0]
        ch1_timedelta = [i-first_time for i in ch1_datetime]
        ch1_time = [datetime(2024, 1, 1, 0, i.seconds//60%60, i.seconds%60, 0) for i in ch1_timedelta]
        return ch1_time, ch1_voltage, ch1_current

    def format_plot(self, ax):
        tick_size = 18
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%M:%S'))
        ax.tick_params(axis='x', labelsize=tick_size, colors='black')  # Set tick size and color here
        ax.tick_params(axis='y', labelsize=tick_size, colors='black')  # Set tick size and color here

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(f"Error: You need to supply a config file for this test as the argument! You had {len(sys.argv)-1} arguments!")
    if (len(sys.argv) == 2):
        LDOmeasure(sys.argv[1])
    elif (len(sys.argv) == 3):
        LDOmeasure(sys.argv[1], sys.argv[2])
    else:
        sys.exit(f"Error: You need to supply a config file and optional test name for this program, 2 arguments max. You supplied {sys.argv}, which is {len(sys.argv)-1} arguments")
